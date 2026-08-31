#!/usr/bin/env python3
"""
BUILD ITEM-LIST.XLSX — every item in the game as a filterable workbook.

    node tools/build-item-list.mjs      # 1. classify + export (source of truth)
    python3 tools/build-item-workbook.py   # 2. build the workbook
    # …or just:  tools/build-item-docs.sh

Reads tools/items-export.json, which the Node generator writes from items.js.
That means the OUTCOME GROUPING here is literally the same table the HTML page
uses — one taxonomy, two renderings, no chance of the two disagreeing.

Sheets
  Summary       counts by class and by outcome, as live formulas over All Items
                (ACTIVE items only — retired items don't count here)
  All Items     one row per ACTIVE item, AutoFilter + frozen header — the
                working sheet
  By Outcome    the same rows sorted outcome-first, with a blank band between
                groups, for reading rather than filtering
  Inactive      the RETIRED items (`inactive: true`) — out of every pool, but
                still registered so old saves keep resolving them
  Legend        what each column means and how to group/filter
"""

import json
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
EXPORT = ROOT / "tools" / "items-export.json"
OUT = ROOT / "item-list.xlsx"

if not EXPORT.exists():
    sys.exit("tools/items-export.json missing — run `node tools/build-item-list.mjs` first")

data = json.loads(EXPORT.read_text())
rows, version = data["rows"], data["version"]
# RETIRED (inactive: true) items leave the main tabs and the Summary — they
# get their own sheet, the HTML page's "Retired" section's twin.
active = [r for r in rows if not r["inactive"]]
retired = [r for r in rows if r["inactive"]]

# ── House look ──────────────────────────────────────────────────────────────
FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3B2C")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
ID_FONT = Font(name=FONT, size=9, color="7F7F7F")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
NOTE_FONT = Font(name=FONT, size=9, color="7F7F7F", italic=True)
GROUP_FONT = Font(name=FONT, bold=True, size=11, color="1F3B2C")
FLAG_FONT = Font(name=FONT, size=10, bold=True, color="C00000")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TIER_FILL = {
    "common": PatternFill("solid", fgColor="F2F2F2"),
    "uncommon": PatternFill("solid", fgColor="DDEBF7"),
    "rare": PatternFill("solid", fgColor="FFF2CC"),
}

COLUMNS = [
    ("Class", "cls", 16),
    ("Outcome", "outcome", 22),
    ("Outcome detail", "outcomeDetail", 30),
    ("Also pays", "outcomeAlso", 22),
    ("Item", "label", 22),
    ("id", "id", 20),
    ("Rarity", "tier", 11),
    ("Cost", "price", 7),
    ("Unlocked at start", "startFlag", 9),
    ("Unlock type", "unlockType", 11),
    ("Unlock stat", "unlockStat", 20),
    ("Unlock count", "unlockCount", 8),
    ("Suit limit", "suits", 9),
    ("In-game help text", "description", 78),
    ("Effect key", "effectKey", 20),
    ("Lint", "flags", 14),
]


def cell_value(row, key):
    if key == "startFlag":
        return "no" if row["gated"] else "yes"
    if key == "outcomeAlso":
        return ", ".join(row["outcomeAlso"])
    if key == "price":
        p = row["price"]
        if row["cursed"]:
            return None          # curses are never sold
        return p
    if key == "unlockCount":
        return row["unlockCount"]
    return row.get(key, "")


def write_table(ws, table_rows, start_row=1, group_bands=False):
    """Header + body. Returns the last written row."""
    for c, (title, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=start_row, column=c, value=title)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[start_row].height = 26

    r = start_row
    last_group = None
    for row in table_rows:
        if group_bands and row["outcome"] != last_group:
            if last_group is not None:
                r += 1                                   # a blank band between groups
            last_group = row["outcome"]
            r += 1
            n = sum(1 for x in table_rows if x["outcome"] == row["outcome"])
            gc = ws.cell(row=r, column=1, value=f"{row['outcome']}  ({n})")
            gc.font = GROUP_FONT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLUMNS))
        r += 1
        for c, (_title, key, _w) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=cell_value(row, key))
            cell.font = ID_FONT if key in ("id", "effectKey") else BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "description"))
            if key == "tier" and row["tier"] in TIER_FILL:
                cell.fill = TIER_FILL[row["tier"]]
            if key == "flags" and row["flags"]:
                cell.font = FLAG_FONT
    return r


wb = Workbook()

# ── Sheet 1: Summary ────────────────────────────────────────────────────────
s = wb.active
s.title = "Summary"
s["A1"] = f"NINELIVES — item pool ({version})"
s["A1"].font = TITLE_FONT
s["A2"] = ("Generated from items.js by tools/build-item-list.mjs + tools/build-item-workbook.py. "
           "Every count below is a live COUNTIF over the 'All Items' sheet, so edits there re-total. "
           f"Counts cover ACTIVE items only — {len(retired)} retired items live on the 'Inactive' sheet.")
s["A2"].font = NOTE_FONT

classes = []
for row in active:
    if row["cls"] not in classes:
        classes.append(row["cls"])
cls_order = {name: i for i, name in enumerate(classes)}
outcomes = sorted({r["outcome"] for r in active},
                  key=lambda n: (n.startswith("Curse"), n))

n_rows = len(active)
CLS_COL = "$A$2:$A$%d" % (n_rows + 1)      # on 'All Items'
OUT_COL = "$B$2:$B$%d" % (n_rows + 1)
TIER_COL = "$G$2:$G$%d" % (n_rows + 1)
START_COL = "$I$2:$I$%d" % (n_rows + 1)

r = 4
s.cell(row=r, column=1, value="BY CLASS").font = Font(name=FONT, bold=True, size=11)
r += 1
for c, h in enumerate(["Class", "Total", "common", "uncommon", "rare", "starting", "gated"], start=1):
    cell = s.cell(row=r, column=c, value=h)
    cell.font, cell.fill, cell.border = HEAD_FONT, HEAD_FILL, BORDER
class_first = r + 1
for i, name in enumerate(classes):
    rr = class_first + i
    s.cell(row=rr, column=1, value=name).font = BODY_FONT
    s.cell(row=rr, column=2, value=f"=COUNTIF('All Items'!{CLS_COL},$A{rr})")
    for j, tier in enumerate(["common", "uncommon", "rare"]):
        s.cell(row=rr, column=3 + j,
               value=f"=COUNTIFS('All Items'!{CLS_COL},$A{rr},'All Items'!{TIER_COL},\"{tier}\")")
    s.cell(row=rr, column=6,
           value=f"=COUNTIFS('All Items'!{CLS_COL},$A{rr},'All Items'!{START_COL},\"yes\")")
    s.cell(row=rr, column=7,
           value=f"=COUNTIFS('All Items'!{CLS_COL},$A{rr},'All Items'!{START_COL},\"no\")")
    for c in range(1, 8):
        s.cell(row=rr, column=c).font = BODY_FONT
        s.cell(row=rr, column=c).border = BORDER
class_last = class_first + len(classes) - 1
tot = class_last + 1
s.cell(row=tot, column=1, value="All items").font = Font(name=FONT, bold=True, size=10)
for c in range(2, 8):
    s.cell(row=tot, column=c,
           value=f"=SUM({get_column_letter(c)}{class_first}:{get_column_letter(c)}{class_last})")
    s.cell(row=tot, column=c).font = Font(name=FONT, bold=True, size=10)
    s.cell(row=tot, column=c).border = BORDER
s.cell(row=tot, column=1).border = BORDER

# Outcome block — the grouping the user asked for, counted per class.
r = tot + 3
s.cell(row=r, column=1, value="BY OUTCOME  (what the item pays you in)").font = Font(name=FONT, bold=True, size=11)
r += 1
head = ["Outcome"] + classes + ["Total"]
for c, h in enumerate(head, start=1):
    cell = s.cell(row=r, column=c, value=h)
    cell.font, cell.fill, cell.border = HEAD_FONT, HEAD_FILL, BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center")
out_first = r + 1
for i, name in enumerate(outcomes):
    rr = out_first + i
    s.cell(row=rr, column=1, value=name).font = BODY_FONT
    for j, cname in enumerate(classes):
        s.cell(row=rr, column=2 + j,
               value=f"=COUNTIFS('All Items'!{OUT_COL},$A{rr},'All Items'!{CLS_COL},\"{cname}\")")
    s.cell(row=rr, column=2 + len(classes),
           value=f"=COUNTIF('All Items'!{OUT_COL},$A{rr})")
    for c in range(1, len(head) + 1):
        s.cell(row=rr, column=c).font = BODY_FONT
        s.cell(row=rr, column=c).border = BORDER
out_last = out_first + len(outcomes) - 1
otot = out_last + 1
s.cell(row=otot, column=1, value="All items").font = Font(name=FONT, bold=True, size=10)
for c in range(2, len(head) + 1):
    s.cell(row=otot, column=c,
           value=f"=SUM({get_column_letter(c)}{out_first}:{get_column_letter(c)}{out_last})")
    s.cell(row=otot, column=c).font = Font(name=FONT, bold=True, size=10)
    s.cell(row=otot, column=c).border = BORDER
s.cell(row=otot, column=1).border = BORDER

s.column_dimensions["A"].width = 30
for c in range(2, len(head) + 1):
    s.column_dimensions[get_column_letter(c)].width = 13

note_r = otot + 2
s.cell(row=note_r, column=1,
       value=("An item can pay in more than one currency; these counts use its PRIMARY outcome. "
              "The 'Also pays' column on 'All Items' carries the secondary ones."))
s.cell(row=note_r, column=1).font = NOTE_FONT

# ── Sheet 2: All Items (ACTIVE only) ────────────────────────────────────────
ws = wb.create_sheet("All Items")
flat = sorted(active, key=lambda x: (cls_order[x["cls"]],
                                     {"common": 0, "uncommon": 1, "rare": 2}.get(x["tier"], 9),
                                     x["label"]))
last = write_table(ws, flat)
ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"

# ── Sheet 3: By Outcome (ACTIVE only) ───────────────────────────────────────
ws2 = wb.create_sheet("By Outcome")
by_outcome = sorted(active, key=lambda x: (outcomes.index(x["outcome"]),
                                           cls_order[x["cls"]],
                                           x["label"]))
last2 = write_table(ws2, by_outcome, group_bands=True)
ws2.freeze_panes = "A2"

# ── Sheet 4: Inactive (the RETIRED items) ───────────────────────────────────
wsi = wb.create_sheet("Inactive")
wsi["A1"] = ("RETIRED items (`inactive: true`) — out of EVERY acquisition pool, but still registered: "
             "an old save's copy keeps resolving and firing, and the Collection shows them greyed. "
             "None of these count on the Summary.")
wsi["A1"].font = NOTE_FONT
wsi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
wsi.row_dimensions[1].height = 24
flat_inactive = sorted(retired, key=lambda x: (cls_order.get(x["cls"], 99),
                                               {"common": 0, "uncommon": 1, "rare": 2}.get(x["tier"], 9),
                                               x["label"]))
lasti = write_table(wsi, flat_inactive, start_row=2)
wsi.freeze_panes = "E3"
wsi.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{lasti}"

# ── Sheet 5: Legend ─────────────────────────────────────────────────────────
lg = wb.create_sheet("Legend")
lg["A1"] = "How to read and group this workbook"
lg["A1"].font = TITLE_FONT
legend = [
    ("Sheet: All Items", "One row per ACTIVE item. The header row has AutoFilter — click any header arrow to "
                         "filter. To group by outcome, filter or sort on the 'Outcome' column; to group by what a "
                         "class offers, filter 'Class' too. Retired items are not here — see the Inactive sheet."),
    ("Sheet: By Outcome", "The same ACTIVE rows pre-sorted into outcome groups with a heading and a count per "
                          "group — for reading straight through rather than filtering."),
    ("Sheet: Inactive", "The RETIRED items (`inactive: true`): out of every acquisition pool, but still registered "
                        "so old saves keep resolving and firing them, and the Collection shows them greyed. "
                        "They count nowhere on the Summary."),
    ("Sheet: Summary", "Counts by class and by outcome, ACTIVE items only. Every number is a live "
                       "COUNTIF/COUNTIFS over 'All Items', so it re-totals if rows are filtered out and "
                       "re-added, or edited."),
    ("", ""),
    ("Outcome", "What the item actually pays you in — its headline payoff. Derived from the item's stable effect "
                "id (never its name or prose), from one taxonomy shared with item-list.html."),
    ("Outcome detail", "The specific shape of that payoff — when it fires, what it keys on."),
    ("Also pays", "Secondary currencies. Kamikaze's primary outcome is Peek; it 'also pays' Pile Destruction."),
    ("Class", "Stickers · Pillars · Bases · Same-Powers · Cursed Stickers · Packs."),
    ("Rarity", "How often the store OFFERS it (common/uncommon/rare) — independent of price. Shaded per tier."),
    ("Cost", "Store price in coins. Blank for cursed stickers, which are never sold."),
    ("Unlocked at start", "yes = in every roll pool from a fresh save. no = gated behind the unlock columns."),
    ("Unlock type / stat / count", "The gate: the lifetime stat and the threshold that opens the item."),
    ("Suit limit", "Stickers only — the suits this sticker may be applied to. Blank = any suit."),
    ("In-game help text", "The EXACT string the store, collection and hold-help show. {rank}/{suit} tokens are "
                          "substituted live before the player sees them."),
    ("Effect key", "The engine's stable behavior/effect id — what the taxonomy is keyed on, and what to grep for "
                   "in the Swift source."),
    ("Lint", "Blank is good. TEMPLATE = carries a live-substituted token (expected). EMPTY / UNWIRED / "
             "\"ROLLED\" / UNCLASSIFIED are problems — see item-list.html for the detail."),
    ("", ""),
    ("Regenerate", "tools/build-item-docs.sh  (or: node tools/build-item-list.mjs && python3 tools/build-item-workbook.py)"),
    ("Source of truth", "items.js at the repo root. Never edit this workbook to change the game — edit items.js "
                        "and regenerate."),
]
for i, (k, v) in enumerate(legend, start=3):
    lg.cell(row=i, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
    c = lg.cell(row=i, column=2, value=v)
    c.font = BODY_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
lg.column_dimensions["A"].width = 28
lg.column_dimensions["B"].width = 105

wb.save(OUT)
print(f"item-list.xlsx written — {len(active)} active + {len(retired)} retired items, "
      f"{len(outcomes)} outcome groups ({version})")
for name in outcomes:
    print(f"  {sum(1 for r in active if r['outcome'] == name):3}  {name}")
